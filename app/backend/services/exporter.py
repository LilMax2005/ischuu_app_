from __future__ import annotations

from pathlib import Path
from tempfile import NamedTemporaryFile

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def export_orders_to_excel(orders: list[dict]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos"

    headers = [
        "ID Pedido", "Fecha", "Usuario", "Estado", "Estado Pago", "Productos",
        "Subtotal", "Envío", "Descuento", "Total", "Puntos Ganados", "Buy Order",
    ]
    ws.append(headers)

    fill = PatternFill("solid", fgColor="EC4899")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

    for order in orders:
        items = order.get("items", []) or []
        items_text = ", ".join(f"{i.get('name','Producto')} x{i.get('quantity',1)}" for i in items)
        ws.append([
            str(order.get("_id", "")), order.get("created_at", ""), order.get("user_email", order.get("user_id", "")),
            order.get("status", ""), order.get("payment_status", ""), items_text,
            int(order.get("subtotal", 0)), int(order.get("shipping", 0)), int(order.get("discount", 0)),
            int(order.get("total", 0)), int(order.get("points_earned", 0)), order.get("buy_order", ""),
        ])

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        width = min(max(len(str(c.value or "")) for c in col) + 2, 45)
        ws.column_dimensions[col_letter].width = width

    tmp = NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.close()
    path = Path(tmp.name)
    wb.save(path)
    return path
